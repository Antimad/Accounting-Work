import calendar
import datetime
import os
import re
import shelve
import string
import sys
import numpy as np
import pandas as pd
from PyQt5 import QtCore
from PyQt5.QtGui import QRegExpValidator, QIcon

from PyQt5.QtWidgets import (QApplication, QFileDialog, QWidget, QPushButton, QGridLayout, QLabel,
                             QComboBox, QSizePolicy, QRadioButton, QLineEdit)
from openpyxl import load_workbook, Workbook, worksheet
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

files = ['Tender',
         'EMP Sale',
         'No Tax',
         'Purchased GC',
         'Redeemed GC',
         'CM Report',
         'FOLDER CONTAINING ALL FILES...']

Date = datetime.datetime.now()
coordinates = [(x, y) for x in range(len(files)) for y in range(1)]
FileLocations = {'File Name': [], 'Location': []}
ReportTime = {'Year': [], 'Month': [Date.month]}
book = []
shelf_files = 'shelve.out'
my_shelf = shelve.open(shelf_files)

icon = r'C:\\Users\\Uchenna\\Documents\\Python\\Icon\\Nadeau.png'


def on_month_choice(selection):
    ReportTime['Month'].append(selection + 1)


class FileSelector(QWidget):
    NewFile = '%s Report 2020' % calendar.month_name[Date.month]

    def __init__(self):
        # noinspection PyArgumentList
        super(FileSelector, self).__init__()
        self.option = None
        self.title = 'Daily Summary Report'
        self.setWindowIcon(QIcon(icon))
        self.selection = ''
        self.FileName = QLineEdit(self)
        self.change_btn = QPushButton(self)
        reg_ex = QtCore.QRegExp("[a-z-A-Z_0-9_. ]+")
        input_validator = QRegExpValidator(reg_ex, self.FileName)
        self.FileName.setValidator(input_validator)
        self.left = 900
        self.top = 500
        self.width = 450
        self.height = 200
        self.grid_layout = QGridLayout()
        self.setLayout(self.grid_layout)
        self.grid_layout.setHorizontalSpacing(40)
        self.greeting()
        self.month_options()
        self.year_options()
        self.file_name_entrance()
        self.layout().setSpacing(25)

    def greeting(self):
        self.setGeometry(self.left, self.top, self.width, self.height)
        self.setWindowTitle(self.title)

        hello = QLabel('Folder:', self)
        # hello.move(QtCore.Qt.AlignCenter, 50)
        hello.setStyleSheet('font:18pt "Segoe"; font-weight:200')
        self.grid_layout.addWidget(hello, 1, 0)

        btn = QPushButton('Search', self)
        btn.clicked.connect(self.get_directory)
        # btn.move(QtCore.Qt.AlignCenter, 10)
        btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        btn.setStyleSheet('font:10pt "Segoe"')
        self.grid_layout.addWidget(btn, 2, 0)
        # noinspection PyTypeChecker
        btn.clicked.connect(self.close)

    def month_options(self):
        working_month = QLabel('Month:', self)
        working_month.setStyleSheet('font:18pt "Segoe"; font-weight:200')
        self.grid_layout.addWidget(working_month, 1, 2)
        combo_box = QComboBox(self)
        combo_box.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.grid_layout.addWidget(combo_box, 2, 2)

        for mth in range(1, 13):
            combo_box.addItem(calendar.month_name[mth])

        # combo_box.move(QtCore.Qt.AlignCenter+350, 150)
        combo_box.setCurrentIndex(Date.month - 1)

        combo_box.currentIndexChanged.connect(on_month_choice)

    def year_options(self):
        current_year = Date.year

        year = QRadioButton(str(current_year))
        year.option = str(current_year)
        year.toggled.connect(self.on_year_choice)
        year.setChecked(True)
        self.grid_layout.addWidget(year, 3, 1)

        year_m1 = QRadioButton(str(current_year - 1))
        year_m1.option = str(current_year - 1)
        year_m1.toggled.connect(self.on_year_choice)
        self.grid_layout.addWidget(year_m1, 3, 0)

        year_p1 = QRadioButton(str(current_year + 1))
        year_p1.option = str(current_year + 1)
        year_p1.toggled.connect(self.on_year_choice)
        self.grid_layout.addWidget(year_p1, 3, 2)

    def on_year_choice(self):
        choice = self.sender()
        # noinspection PyUnresolvedReferences
        if choice.isChecked():
            ReportTime['Year'].append(int(choice.option))
            print(choice.option)

    def file_name_entrance(self):
        self.FileName.setReadOnly(True)
        self.grid_layout.addWidget(self.FileName, 0, 1, 1, 2)

        self.change_btn.setText('Change')
        self.change_btn.clicked.connect(self.change_requested)
        self.grid_layout.addWidget(self.change_btn, 0, 0)
        try:
            self.FileName.setPlaceholderText(str(my_shelf['current_file']) +
                                             ' will be used')
            working_filename = my_shelf['current_file']
            work_book = load_workbook(working_filename + '.xlsx')
        except KeyError:  # This means that there is no report on the shelf
            self.FileName.setPlaceholderText('What is the name of the report?')
            self.change_requested()
            name = FileSelector.NewFile
            working_filename = name
            my_shelf['current_file'] = working_filename
            try:
                work_book = load_workbook(working_filename + '.xlsx')
                # print('Found the file')
            except FileNotFoundError:
                work_book = Workbook()
                # print('Creating a new file')
        except FileNotFoundError:  # This error is because though a filename is on the shelf, but it isn't in the folder
            working_filename = str(my_shelf['current_file'])
            self.FileName.setPlaceholderText(working_filename + " wasn't found")
            name = FileSelector.NewFile
            working_filename = name
            print(working_filename)
            my_shelf['current_file'] = working_filename
            try:
                work_book = load_workbook(working_filename + '.xlsx')
                print('Found the file')
                self.FileName.setPlaceholderText(working_filename + " wasn't found")
            except FileNotFoundError:
                work_book = Workbook()
                print('Creating a new file')
        book.append(work_book)

    def change_requested(self):
        self.FileName.setReadOnly(False)
        self.grid_layout.removeWidget(self.change_btn)
        self.change_btn.deleteLater()
        self.change_btn = None
        save_btn = QPushButton('Save', self)
        self.grid_layout.addWidget(save_btn, 0, 0)
        save_btn.clicked.connect(self.save_file)

    def save_file(self):
        if not self.FileName.text().isspace() and self.FileName.text() != '':
            print('no blanks here', self.FileName.text())
            FileSelector.NewFile = self.FileName.text()
            my_shelf['current_file'] = FileSelector.NewFile
            self.FileName.setText(FileSelector.NewFile + ' Saved!')
            work_book = Workbook()
            book.append(work_book)
            print('Clicked save and should create new workbook')

    def search_file(self):
        options = QFileDialog.Options()
        # noinspection PyCallByClass
        find_file, _ = QFileDialog.getOpenFileName(self, 'Purchase Order', '',
                                                   'Excel Files (*.xlsx *xls)',
                                                   options=options)
        FileLocations['Location'].append(find_file)

    def get_directory(self):
        dialog = QFileDialog()
        folder_path = dialog.getExistingDirectory(self, '', 'Select Folder with ALL 6 Files')
        FileLocations['File Name'].append('Directory')
        FileLocations['Location'].append(folder_path)
        dialog.setEnabled(False)
        self.close()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    window = FileSelector()
    window.show()
    app.exec_()

my_shelf.sync()
wb = book[-1]
Month = ReportTime['Month'][-1]
Year = ReportTime['Year'][-1]

if Month != datetime.datetime.now().month:
    filename = '%s Report 2020' % calendar.month_name[Month]
    wb = load_workbook(filename + '.xlsx')
    print('should be working with the old one')
else:
    filename = '%s Report 2020' % calendar.month_name[Date.month]

work_sheet = wb.active
work_sheet.cell_addition = calendar.month_name[Month] + ' 2020'

Cash_Columns = {
    'Nadeau Reports': 'D',
    'Per Statements': 'E',
    'Difference': 'F'
}

Check_Columns = {
    'Nadeau Reports': 'H',
    'Per Statements': 'I',
    'Difference': 'J'
}

VisaMCD_Columns = {
    'Nadeau Reports': 'L',
    'Per Statements': 'M',
    'Difference': 'N'
}

Amex_Columns = {
    'Nadeau Reports': 'P',
    'Per Statements': 'Q',
    'Difference': 'R'
}

Square_Columns = {
    'Nadeau Reports': 'T',
    'Per Statements': 'U',
    'Difference': 'V'
}

Totals_Columns = {
    'Credit Card': 'W',
    'Grand Total': 'X',
    'GC Used': 'Y',
    'Tax Included': 'Z'
}

Trailing_Columns = {
    'Verification': 'AA',
    'GC bought': 'AB',
    'Tax Exempt': 'AC',
    'PPS2': 'AD',
    'Total Employee_1': 'AE',
    'Difference_1': 'AG',
    'PP1': 'AH',
    'Total Employee_2': 'AI',
    'Difference_2': 'AK'
}

POFile = FileLocations['Location'][0]


def missing_tables(df):
    try:
        df
    except NameError:
        df = pd.DataFrame({'A': []})
    return df


FileLocations = pd.DataFrame(FileLocations)
FileLocations = FileLocations.set_index('File Name')
FileLocations = FileLocations.drop_duplicates()

Tendered = EmpDisc = TaxFreeSales = RedeemedGC = PurchasedGC = CreditMemo = Tax_Exempt = None


def obedience(main_report):
    n_tendered = pd.read_excel(main_report, skiprows=5)
    n_tendered.drop([0, 1, 2], inplace=True)
    n_tendered.set_index(['Unnamed: 0'], inplace=True)
    n_tendered.index = pd.Series(n_tendered.index).fillna(method='ffill')
    n_tendered.dropna(axis='columns', how='all', inplace=True)
    n_tendered.dropna(axis='index', subset=['Unnamed: 7'], inplace=True)

    """
    Comment on dropping all NaN in n_tendered columns. 
    
    I think it is best to not do it given the small cases where there are a singular sale in a payment category that is
    also Tax Free. 
    This is a very unique situation, but still possible where the tax column is empty resulting in the remaining columns
    after the occurrence to be out of place.
    
    """

    current_column = n_tendered.columns.to_list()

    amex_dups = v_mc_d = 1
    for i, name in enumerate(current_column):
        # Commission = i + 1
        # ItemTax = i + 2
        if name == 'DOLLARS':
            current_column[i] = 'Cash'
            # current_column[Commission] = 'Cash Commission'
            # current_column[ItemTax] = 'Cash Taxed'

        """
        if name == 'Check':
            current_column[Commission] = 'Check Commission'
            current_column[ItemTax] = 'Check Taxed'
        """

        if name == 'AMEX.1':
            amex_dups += 1
            current_column[i] = 'AMEX_%s' % amex_dups
            # current_column[Commission] = 'AMEX Commission_%s' % amex_dups
            # current_column[ItemTax] = 'AMEX Taxed_%s' % amex_dups

        if name == 'V/MC/D.1':
            v_mc_d += 1
            current_column[i] = 'VisaMCD_%s' % v_mc_d
            # current_column[Commission] = 'VisaMCD Commission_%s' % v_mc_d
            # current_column[ItemTax] = 'VisaMCD Taxed_%s' % v_mc_d

        if name == 'V/MC/D':
            current_column[i] = 'VisaMCD_1'
            # current_column[Commission] = 'VisaMCD Commission'
            # current_column[ItemTax] = 'VisaMCD Taxed'

        if name == 'AMEX':
            current_column[i] = 'AMEX_1'
            # current_column[Commission] = 'AMEX Commission'
            # current_column[ItemTax] = 'AMEX Taxed'

        if name == 'SQUARE':
            current_column[i] = 'Square'
            # current_column[Commission] = 'Square Commission'
            # current_column[ItemTax] = 'Square Taxed'

        if name == 'Gift Card':
            current_column[i] = 'GCTotal'
            # current_column[Commission] = 'GCTotal Commission'
            # current_column[ItemTax] = 'GCTotal Taxed'

        if name == 'Store Credit':
            current_column[i] = 'SCTotal'
            # current_column[Commission] = 'SCTotal Commission'
            # current_column[ItemTax] = 'SCTotal Taxed'

    current_column[-3] = 'GTotal'
    current_column[-2] = 'GTotal Commission'
    current_column[-1] = 'GTotal Taxed'
    current_column[0] = 'Date'

    n_tendered.columns = current_column

    if amex_dups == 2:
        n_tendered['AMEX'] = n_tendered['AMEX_1'].add(n_tendered['AMEX_2'], fill_value=0)
    else:
        n_tendered.rename(columns={'AMEX_1': 'AMEX'}, inplace=True)

    if v_mc_d == 2:
        n_tendered['VisaMCD'] = n_tendered['VisaMCD_1'].add(n_tendered['VisaMCD_2'], fill_value=0)
    else:
        n_tendered.rename(columns={'VisaMCD_1': 'VisaMCD'}, inplace=True)

    return n_tendered


if FileLocations['Location']['Directory']:
    for index, folder_files in enumerate(os.listdir(FileLocations['Location']['Directory'])):
        for report in range(len(files)):
            if re.search(files[report], folder_files, re.IGNORECASE):
                if files[report] == 'Tender':
                    TenderReport = FileLocations['Location']['Directory'] + '/' + folder_files
                    while True:
                        try:
                            Tendered = obedience(TenderReport)
                            break
                        except NameError:
                            print("The Tender Report must be named 'Tender' only")

                if files[report] == 'EMP Sale':
                    EmpSales = FileLocations['Location']['Directory'] + '/' + folder_files

                    while True:
                        try:
                            EmpDisc = pd.read_excel(EmpSales)
                            EmpDisc = EmpDisc.set_index(['Store Name'])
                            break
                        except NameError:
                            print("The Employee Sales Report must be named 'Employee Sale' only")

                if files[report] == 'No Tax':
                    TaxFreeSales = FileLocations['Location']['Directory'] + '/' + folder_files

                    while True:
                        try:
                            Tax_Exempt = pd.read_excel(TaxFreeSales)
                            Tax_Exempt = Tax_Exempt.set_index(['Store Name'])
                            break
                        except NameError:
                            print("The Tax Exemption Report must be named 'Tax Free Sale'")
                            continue

                if files[report] == 'Purchased GC':
                    GC_Sales = FileLocations['Location']['Directory'] + '/' + folder_files

                    while True:
                        try:
                            PurchasedGC = pd.read_excel(GC_Sales)
                            PurchasedGC = PurchasedGC.set_index('Store Name')
                            break
                        except NameError:
                            print("The Purchased GCs should be named 'Purchased GC' only")
                            continue

                if files[report] == 'Redeemed GC':
                    GC_Used = FileLocations['Location']['Directory'] + '/' + folder_files

                    while True:
                        try:
                            RedeemedGC = pd.read_excel(GC_Used)
                            RedeemedGC = RedeemedGC.set_index('Store Name')
                            break
                        except NameError:
                            print("The Redeemed GCs should be named 'Redeemed GC' only")
                            continue

                if files[report] == 'CM Report':
                    CM_Sales_Issuance = FileLocations['Location']['Directory'] + '/' + folder_files

                    while True:
                        try:
                            CreditMemo = pd.read_excel(CM_Sales_Issuance)
                            break
                        except NameError:
                            print("The Credit Memos should be named 'CM Report' only")
                            continue

TaxRate = 'Tax Rate.xlsx'
# PurchasedGC = pd.DataFrame({'A': []})

EmptyDF = pd.DataFrame({'A': []})
# data_bases = [Tendered, EmpDisc, TaxFreeSales, RedeemedGC, PurchasedGC, CreditMemo]

if EmpDisc is None:
    EmpDisc = EmptyDF
    print('No Employee Discount file found')

if TaxFreeSales is None:
    Tax_Exempt = EmptyDF
    print('No Tax Exempt file found')

if RedeemedGC is None:
    RedeemedGC = EmptyDF
    print('No Redeemed GC file found')

if PurchasedGC is None:
    PurchasedGC = EmptyDF
    print('No Purchased GC file found')

if CreditMemo is None:
    CreditMemo = EmptyDF
    print('No CreditMemo file found')

Tax = pd.read_excel(TaxRate)
Tax = Tax.set_index(['Headquarters'])

try:
    CreditMemo['Invoice #'] = CreditMemo['Invoice #'].fillna(0)
    CreditMemo = CreditMemo.astype({'Invoice #': 'int'})
    CreditMemo = CreditMemo.set_index('Invoice #')
except ValueError:
    invoices = []
    for CMs in range(len(CreditMemo)):
        try:
            invoices.append(CreditMemo['Invoice #'][CMs].split()[0])
        except AttributeError:
            invoices.append(CreditMemo['Invoice #'][CMs])
    CreditMemo['Invoice #'] = invoices
    CreditMemo = CreditMemo.set_index('Invoice #')
except KeyError:
    pass

Locations = ['Alexandria', 'Asheville', 'Austin', 'Baton Rouge', 'Birmingham', 'Boston', 'Buckhead', 'Charleston',
             'Charlotte', 'Chattanooga', 'Chicago', 'Cincinnati', 'Columbia', 'Dallas', 'Detroit', 'Fort Worth',
             'Houston', 'Huntsville', 'Indianapolis', 'Kansas City', 'Knoxville', 'Little Rock', 'Los Angeles',
             'Louisville', 'Marietta', 'Memphis', 'Miami', 'Minneapolis', 'Nashville', 'New Orleans',
             'Orlando', 'Paramus', 'Philadelphia', 'Pittsburgh', 'Portland', 'Raleigh', 'Savannah', 'Tampa']

Locations_Key = {
    'Birmingham': 'AL (HOMEWOOD)',
    'Huntsville': 'AL1 (HUNTSVILLE)',
    # 'Little Rock': 'AR (LITTLE ROCK)',
    'The Factory': 'CA2 (LOS ANGELES)',
    # 'Sherman Oaks': 'CA3 (SHERMAN OAKS)',
    'Miami': 'FL (MIAMI)',
    # 'Orlando': 'FL3 (ORLANDO)',
    'Tampa': 'FL4 (TAMPA)',
    'Marietta': 'GA1 (MARIETTA)',
    'Savannah': 'GA2 (SAVANNAH)',
    'Buckhead': 'GA3 (BUCKHEAD)',
    'Chicago': 'IL (CHICAGO)',
    'Indianapolis': 'IN (INDIANAPOLIS)',
    # 'Louisville': 'KY (LOUISVILLE)',
    'Kansas City': 'KS (KANSAS)',
    'New Orleans': 'LA (NOLA)',
    'Baton Rouge': 'LA1 (BATON ROUGE)',
    'Boston': 'MA (BOSTON)',
    'Detroit': 'MI (DETROIT)',
    'Minneapolis': 'MN (MINNEAPOLIS)',
    'Charlotte': 'NC (CHARLOTTE)',
    'Raleigh': 'NC1 (RALEIGH)',
    'Asheville': 'NC2 (ASHEVILLE)',
    # 'Paramus': 'NJ1 (PARAMUS)',
    'Cincinnati': 'OH (CINCINNATI)',
    'Portland': 'OR1 (WSM)',
    # 'Philadelphia': 'PA (PHILADELPHIA)',
    'Pittsburgh': 'PA1 (PITTSBURGH)',
    'Mount Pleasant': 'SC2 (MTP)',
    'Columbia': 'SC5 (COLUMBIA)',
    'Nashville': 'TN (NASHVILLE)',
    'Memphis': 'TN1 (MEMPHIS)',
    'Knoxville': 'TN2 (KNOXVILLE)',
    'Chattanooga': 'TN3 (CHATTANOOGA)',
    'Dallas': 'TX (DALLAS)',
    'Houston': 'TX1 (HOUSTON)',
    'Austin': 'TX2 (AUSTIN)',
    # 'San Antonio': 'TX3 (SAN ANTONIO)',
    # 'Fort Worth': 'TX4 (FT WORTH)',  # Closes near end of January
    'Alexandria': 'VA (ALEXANDRIA)'
}

RED = 'FF0000'
work_sheet.column_dimensions['A'].width = 25
Bold = Font(bold=True)
BoldRed = Font(bold=True, color=RED)
Normal = Font(bold=False)
Hor_Center = Alignment(horizontal='center', vertical='bottom')
Hor_Left = Alignment(horizontal='left', vertical='bottom')
TextWrap = Alignment(wrap_text=True, horizontal='center', vertical='bottom')
TitleBorder = Border(top=Side(border_style='thin', color='000000'),
                     bottom=Side(border_style='thin', color='000000'))


def cell_addition(text, working_cell, font, new=False, alignment=None, border=None, merge='A1', number_format='General',
                  column_width=False, place=0):
    if new:
        try:
            text = text.iat[place]
        except AttributeError:
            text = text

    work_sheet.merge_cells(merge)
    work_sheet[working_cell].font = font
    work_sheet[working_cell].alignment = alignment
    work_sheet[working_cell] = text
    work_sheet[working_cell].border = border
    work_sheet[working_cell].number_format = number_format
    if column_width:
        col = ''.join([i for i in working_cell if not i.isdigit()])
        work_sheet.column_dimensions[col].width = len(work_sheet[working_cell].value) + 2


def data_frame_try_catch(df, group, location, place):
    try:
        search = df[group][location].iat[place]
        return search
    except AttributeError:
        search = df[group][location]
        return search
    except KeyError:
        return np.nan


def labels():
    title = work_sheet['A1']
    title.alignment = Hor_Center
    description = work_sheet['A2']
    description.alignment = Hor_Left
    title.font = Font(bold=True)
    description.font = Font(bold=True)

    work_sheet['A1'] = 'Nadeau Corporation'
    work_sheet.merge_cells('A2:B2')
    work_sheet['A2'] = 'Transactions Detail - Month Year'
    work_sheet.column_dimensions['A'].width = 20
    work_sheet.column_dimensions['Z'].width = 20
    work_sheet.column_dimensions['AD'].width = 20
    work_sheet.column_dimensions['B'].width = 20

    work_sheet['A5'] = 'Bank'
    work_sheet['B5'] = 'Date'
    work_sheet['A5'].font = Bold
    work_sheet['B5'].font = Bold

    """
    Fourth(4) row titles are just a general name of each payment bracket
    This row also includes titles that were merged to mimic the original formats the department was used to

    """
    row4_heading = 4
    sub_heading = 5
    cell_addition(text='Cash/Check',
                  merge='{1}{0}:{2}{0}'.format(row4_heading, Cash_Columns['Nadeau Reports'],
                                               Check_Columns['Difference']),
                  working_cell='{1}{0}'.format(row4_heading, Cash_Columns['Nadeau Reports']), font=Bold,
                  alignment=Hor_Center)

    cell_addition(text='CA - Credit Card', merge='{1}{0}:{2}{0}'.format(row4_heading, VisaMCD_Columns['Nadeau Reports'],
                                                                        Amex_Columns['Difference']),
                  working_cell='L4'.format(row4_heading, VisaMCD_Columns['Nadeau Reports']), font=Bold,
                  alignment=Hor_Center)

    cell_addition(text='Total Credit Cards',
                  merge='{2}{0}:{2}{1}'.format(row4_heading, sub_heading, Totals_Columns['Credit Card']),
                  working_cell='{1}{0}'.format(row4_heading, Totals_Columns['Credit Card']), font=Bold,
                  alignment=TextWrap, column_width=True)

    cell_addition(text='GRAND TOTAL SALES',
                  merge='{2}{0}:{2}{1}'.format(row4_heading, sub_heading, Totals_Columns['Grand Total']),
                  working_cell='{1}{0}'.format(row4_heading, Totals_Columns['Grand Total']), font=Bold,
                  alignment=TextWrap, column_width=True)

    cell_addition(text='Total Gift Cards Used',
                  merge='{2}{0}:{2}{1}'.format(row4_heading, sub_heading, Totals_Columns['GC Used']),
                  working_cell='{1}{0}'.format(row4_heading, Totals_Columns['GC Used']), font=Bold, alignment=TextWrap,
                  column_width=True)

    cell_addition(text='TAX INCLUDED',
                  merge='{2}{0}:{2}{1}'.format(row4_heading, sub_heading, Totals_Columns['Tax Included']),
                  working_cell='{1}{0}'.format(row4_heading, Totals_Columns['Tax Included']), font=Bold,
                  alignment=TextWrap, column_width=True)

    cell_addition(text='TAX EXEMPT',
                  merge='{2}{0}:{2}{1}'.format(row4_heading, sub_heading, Trailing_Columns['Tax Exempt']),
                  working_cell='{1}{0}'.format(row4_heading, Trailing_Columns['Tax Exempt']), font=Bold,
                  alignment=TextWrap, column_width=True)

    cell_addition(text='Total Employee',
                  merge='{2}{0}:{2}{1}'.format(row4_heading, sub_heading, Trailing_Columns['Total Employee_1']),
                  working_cell='{1}{0}'.format(row4_heading, Trailing_Columns['Total Employee_1']), font=Bold,
                  alignment=TextWrap, column_width=True)

    cell_addition(text='Total Employee',
                  merge='{2}{0}:{2}{1}'.format(row4_heading, sub_heading, Trailing_Columns['Total Employee_2']),
                  working_cell='{1}{0}'.format(row4_heading, Trailing_Columns['Total Employee_2']), font=Bold,
                  alignment=TextWrap, column_width=True)

    """
    End of Row 4 Titles

    """

    """
    Fifth(5) row titles. 
    These are mainly The Nadeau Reports, Difference, and Per Statements, but also includes trailing columns
    the trailing column's tasks are to monitor the initial columns with the different payment methods

    """
    cell_addition(text='Cash', font=Bold, alignment=Hor_Center,
                  merge='{1}{0}:{2}{0}'.format(sub_heading, Cash_Columns['Nadeau Reports'], Cash_Columns['Difference']),
                  border=TitleBorder, working_cell='{1}{0}'.format(sub_heading, Cash_Columns['Nadeau Reports']))

    cell_addition(text='Check',
                  merge='{1}{0}:{2}{0}'.format(sub_heading, Check_Columns['Nadeau Reports'],
                                               Check_Columns['Difference']),
                  working_cell='{1}{0}'.format(sub_heading, Check_Columns['Nadeau Reports']), border=TitleBorder,
                  alignment=Hor_Center, font=Bold)

    cell_addition(text='Visa/MC/Discover',
                  merge='{1}{0}:{2}{0}'.format(sub_heading, VisaMCD_Columns['Nadeau Reports'],
                                               VisaMCD_Columns['Difference']),
                  working_cell='{1}{0}'.format(sub_heading, VisaMCD_Columns['Nadeau Reports']), font=Bold,
                  border=TitleBorder, alignment=Hor_Center)

    cell_addition(text='Amex',
                  merge='{1}{0}:{2}{0}'.format(sub_heading, Amex_Columns['Nadeau Reports'],
                                               Amex_Columns['Difference']),
                  working_cell='{1}{0}'.format(sub_heading, Amex_Columns['Nadeau Reports']), font=Bold,
                  border=TitleBorder, alignment=Hor_Center)

    cell_addition(text='Square',
                  merge='{1}{0}:{2}{0}'.format(sub_heading, Square_Columns['Nadeau Reports'],
                                               Square_Columns['Difference']),
                  working_cell='{1}{0}'.format(sub_heading, Square_Columns['Nadeau Reports']), font=Bold,
                  border=TitleBorder, alignment=Hor_Center)

    cell_addition(text='GC bought', working_cell='{1}{0}'.format(sub_heading, Trailing_Columns['GC bought']),
                  font=BoldRed, alignment=Hor_Center, column_width=True)

    cell_addition(text='PPS2', working_cell='{1}{0}'.format(sub_heading, Trailing_Columns['PPS2']), font=Bold,
                  alignment=Hor_Center)

    cell_addition(text='Difference', working_cell='{1}{0}'.format(sub_heading, Trailing_Columns['Difference_1']),
                  font=Bold, alignment=Hor_Center, column_width=True)

    cell_addition(text='PP1', working_cell='{1}{0}'.format(sub_heading, Trailing_Columns['PP1']), font=Bold,
                  alignment=Hor_Center)

    cell_addition(text='Difference', working_cell='{1}{0}'.format(sub_heading, Trailing_Columns['Difference_2']),
                  font=Bold, alignment=Hor_Center, column_width=True)

    """
    End of Row 5 Titles

    """

    sub_titles = ['Nadeau Reports', 'Per Statements', 'Difference', ' ']

    letters = string.ascii_uppercase
    letters = letters.replace('ABC', '')
    letters = letters.replace('WXYZ', '')

    subs = 0
    for letter in range(len(letters)):
        cell_addition(text=sub_titles[subs], working_cell=letters[letter] + '6', font=Bold, alignment=Hor_Center,
                      column_width=True)
        if subs >= 3:
            subs = -1
        subs += 1

    work_sheet.freeze_panes = 'C7'


labels()

RowFillSeparator = PatternFill(fill_type='solid', start_color='FFFF0000', end_color='FFFF0000')
RowBorderSeparator = Border(bottom=Side(style='thick'))

Currency = '$#,##0.00_);[Red]($#,##0.00)'
Month_Range = calendar.monthrange(Year, Month)
Locations_Info = []
CellValue = 6
Row = FullMonth = MonthFirst = CurrentTax = 0
for BankIndex, Bank in enumerate(Locations_Key.keys()):
    for FullMonth in range(Month_Range[1]):

        CurrentTax = Tax['tax'][Bank.replace('\n', '')]
        Date = datetime.date(Year, Month, (FullMonth + 1))
        CellValue += 1
        Row = str(CellValue)
        cell_addition(text=Locations_Key[str(Bank)], working_cell=('A' + Row), font=Bold)  # BANK
        cell_addition(text=Date.strftime('%m/%d/%Y'), working_cell=('B' + Row), font=Bold)  # DATE
        cell_addition(text=('={1}{0}-{2}{0}'.format(Row, Cash_Columns['Per Statements'],
                                                    Cash_Columns['Nadeau Reports'])),
                      working_cell=('{1}{0}'.format(Row, Cash_Columns['Difference'])), font=Normal,
                      number_format=Currency)
        cell_addition(text=('={1}{0}-{2}{0}'.format(Row, Check_Columns['Per Statements'],
                                                    Check_Columns['Nadeau Reports'])),
                      working_cell=('{1}{0}'.format(Row, Check_Columns['Difference'])), font=Normal,
                      number_format=Currency)
        cell_addition(text=('={1}{0}-{2}{0}'.format(Row, VisaMCD_Columns['Per Statements'],
                                                    VisaMCD_Columns['Nadeau Reports'])),
                      working_cell=('{1}{0}'.format(Row, VisaMCD_Columns['Difference'])),
                      font=Normal, number_format=Currency)
        cell_addition(text=('={1}{0}-{2}{0}'.format(Row, Amex_Columns['Per Statements'],
                                                    Amex_Columns['Nadeau Reports'])),
                      working_cell=('{1}{0}'.format(Row, Amex_Columns['Difference'])),
                      font=Normal, number_format=Currency)
        cell_addition(text=('={1}{0}-{2}{0}'.format(Row, Square_Columns['Per Statements'],
                                                    Square_Columns['Nadeau Reports'])),
                      working_cell=('{1}{0}'.format(Row, Square_Columns['Difference'])),
                      font=Normal, number_format=Currency)

        cell_addition(text=('={1}{0}/({2}{0}-{3}{0}+{4}{0}-{5}{0})'.format(Row, Totals_Columns['Tax Included'],
                                                                           Totals_Columns['Grand Total'],
                                                                           Totals_Columns['Tax Included'],
                                                                           Trailing_Columns['GC bought'],
                                                                           Trailing_Columns['Tax Exempt'])),
                      working_cell=('{1}{0}'.format(Row, Trailing_Columns['Verification'])),
                      font=Normal, number_format='0.0000')
        cell_addition(text=('={1}{0}+{2}{0}+{3}{0}'.format(Row, VisaMCD_Columns['Nadeau Reports'],
                                                           Amex_Columns['Nadeau Reports'],
                                                           Square_Columns['Nadeau Reports'])),
                      working_cell=Totals_Columns['Credit Card'] + Row, font=Normal,
                      number_format=Currency)
        cell_addition('={1}{0}+{2}{0}+{3}{0}+{4}{0}'.format(Row, Totals_Columns['Credit Card'],
                                                            Cash_Columns['Nadeau Reports'],
                                                            Check_Columns['Nadeau Reports'],
                                                            Totals_Columns['GC Used']),
                      working_cell=Totals_Columns['Grand Total'] + Row, font=Normal, number_format=Currency)
        if Date.day <= 15:
            cell_addition(text=('={1}{0}-{2}{0}-{3}{0}+{4}{0}'.format(Row, Totals_Columns['Grand Total'],
                                                                      Totals_Columns['Tax Included'],
                                                                      Trailing_Columns['Total Employee_2'],
                                                                      Trailing_Columns['GC bought'])),
                          working_cell=Trailing_Columns['PP1'] + Row, font=Normal, number_format=Currency)
        else:
            cell_addition(text=('={1}{0}-{2}{0}-{3}{0}+{4}{0}'.format(Row, Totals_Columns['Grand Total'],
                                                                      Totals_Columns['Tax Included'],
                                                                      Trailing_Columns['Total Employee_1'],
                                                                      Trailing_Columns['GC bought'])),
                          working_cell=Trailing_Columns['PPS2'] + Row, font=Normal, number_format=Currency)
        UnTaxed = 0
        GCsPurchased = 0
        GCsRedeemed = 0
        if Bank in Tendered.index:
            try:
                len(PurchasedGC['Date'][Bank])
                PurGo = True
                # noinspection PyTypeChecker
                for GCs in enumerate(pd.to_datetime(PurchasedGC['Date'][Bank])):
                    if Date == GCs[1]:
                        GCsPurchased -= PurchasedGC['GIVEN'][Bank].iat[GCs[0]]
                        cell_addition(text=GCsPurchased, working_cell=Trailing_Columns['GC bought'] + Row,
                                      font=Normal, number_format=Currency)
            except TypeError:
                PurGo = False
                if pd.to_datetime(PurchasedGC['Date'][Bank]) == Date:
                    GCsPurchased = -PurchasedGC['GIVEN'][Bank]
                    cell_addition(text=GCsPurchased, working_cell=Trailing_Columns['GC bought'] + Row,
                                  font=Normal, number_format=Currency)
            except KeyError:
                pass

            try:
                len(RedeemedGC['Date'][Bank])
                RedGo = True
                # noinspection PyTypeChecker
                for RGC in enumerate(pd.to_datetime(RedeemedGC['Date'][Bank])):
                    if Date == RGC[1]:
                        GCsRedeemed += RedeemedGC['TAKEN'][Bank].iat[RGC[0]]
                        cell_addition(text=GCsRedeemed, working_cell=Totals_Columns['GC Used'] + Row,
                                      font=Normal, number_format=Currency)
            except TypeError:
                RedGo = False
                if Date == pd.to_datetime(RedeemedGC['Date'][Bank]):
                    GCsRedeemed = RedeemedGC['TAKEN'][Bank]
                    cell_addition(text=GCsRedeemed, working_cell=Totals_Columns['GC Used'] + Row,
                                  font=Normal, number_format=Currency)
            except KeyError:
                pass

            for data in range(len(Tendered['Date'][Bank])):
                try:
                    working_date = pd.to_datetime(Tendered['Date'][Bank].iat[data])
                except AttributeError:
                    working_date = pd.to_datetime(Tendered['Date'][Bank])
                if Date == working_date:
                    try:
                        UnTaxed = Memo = 0
                        for TaxFree in range(len(Tax_Exempt['Date'][Bank])):
                            if Date == pd.to_datetime(Tax_Exempt['Date'][Bank].iat[TaxFree]):
                                UnTaxed += Tax_Exempt['Item Subtotal'][Bank].iat[TaxFree]
                                try:
                                    Memo += CreditMemo['AMT'][str(Tax_Exempt['Invoice #'][Bank].iat[TaxFree])]
                                except (KeyError, IndexError):
                                    pass
                        cell_addition(text=UnTaxed - Memo, working_cell=(Trailing_Columns['Tax Exempt'] + Row),
                                      font=Normal, number_format=Currency)
                        UnTaxed = Memo = 0
                    except KeyError:
                        pass
                    except TypeError:
                        """
                        Sometimes there is only one Credit Memo, so we can't use .iat function to count through the CMs
                        """
                        TaxFree = 0
                        if Date == pd.to_datetime(Tax_Exempt['Date'][Bank]):
                            cell_addition(text=(Tax_Exempt['Item Subtotal'][Bank]),
                                          working_cell=(Trailing_Columns['Tax Exempt'] + Row),
                                          font=Normal, number_format=Currency)

                    # Cash Total
                    cell_addition(text=Tendered['Cash'][Bank], working_cell=(Cash_Columns['Nadeau Reports'] + Row),
                                  font=Normal, number_format=Currency, new=True, place=data)
                    # Check Total
                    try:
                        cell_addition(text=Tendered['Check'][Bank], working_cell=Check_Columns['Nadeau Reports'] + Row,
                                      font=Normal, number_format=Currency, new=True, place=data)
                    except KeyError:
                        pass
                    cell_addition(text=Tendered['VisaMCD'][Bank], working_cell=VisaMCD_Columns['Nadeau Reports'] + Row,
                                  font=Normal, number_format=Currency, new=True, place=data)
                    cell_addition(text=Tendered['AMEX'][Bank], working_cell=Amex_Columns['Nadeau Reports'] + Row,
                                  font=Normal, number_format=Currency, new=True, place=data)
                    try:
                        cell_addition(text=Tendered['Square'][Bank],
                                      working_cell=Square_Columns['Nadeau Reports'] + Row,
                                      font=Normal, number_format=Currency, new=True, place=data)
                    except KeyError:
                        pass

                    GCsPurchased = 0
                    GCsRedeemed = 0
                    RedGo = False
                    PurGo = False

                    # ****SC is purposefully being removed from the GTotal & GTotal Taxed****

                    GrandTotal = data_frame_try_catch(df=Tendered, group='GTotal', location=Bank, place=data)

                    GCTotal = data_frame_try_catch(df=Tendered, group='GCTotal', location=Bank, place=data)

                    if not np.isnan(GCTotal):
                        if GCTotal < 0:
                            GrandTotal -= GCTotal

                    SCTotal = data_frame_try_catch(df=Tendered, group='SCTotal', location=Bank, place=data)

                    if np.isnan(SCTotal):
                        GTTaxed = data_frame_try_catch(df=Tendered, group='GTotal Taxed', location=Bank, place=data)
                        cell_addition(text=round(GTTaxed, 2), working_cell=Totals_Columns['Tax Included'] + Row,
                                      font=Normal, number_format=Currency)
                        if Date.day <= 15:
                            # PAY PERIOD
                            # This Try Block is adding the Employee payments
                            try:
                                try:
                                    EmpDayTotal = 0
                                    for discount in range(len(EmpDisc['Date'][Bank])):
                                        EmpDiscDate = data_frame_try_catch(df=EmpDisc, group='Date',
                                                                           location=Bank, place=discount)

                                        if Date.day == pd.to_datetime(EmpDiscDate).day:
                                            EmpDayTotal += data_frame_try_catch(df=EmpDisc, group='Item Subtotal',
                                                                                location=Bank, place=discount)
                                            cell_addition(text=EmpDayTotal,
                                                          working_cell=Trailing_Columns['Total Employee_2'] + Row,
                                                          font=Normal, number_format=Currency)
                                except TypeError:
                                    if Date.day == pd.to_datetime(EmpDisc['Date'][Bank]).day:
                                        cell_addition(text=EmpDisc['Item Subtotal'][Bank],
                                                      working_cell=Trailing_Columns['Total Employee_2'] + Row,
                                                      font=Normal,
                                                      number_format=Currency)
                                    #  print(EmpDisc['Item Subtotal'][Bank], Date.day)
                            except KeyError:
                                pass
                        else:
                            # PAY PERIODS
                            cell_addition(
                                text=('={1}{0}-{2}{0}-{3}{0}+{4}{0}'.format(Row, Totals_Columns['Grand Total'],
                                                                            Totals_Columns['Tax Included'],
                                                                            Trailing_Columns[
                                                                                'Total Employee_1'],
                                                                            Trailing_Columns['GC bought'])),
                                working_cell=Trailing_Columns['PPS2'] + Row,
                                font=Normal, number_format=Currency)
                            # This Try Block adds the Employee payments
                            try:
                                try:
                                    EmpDayTotal = 0
                                    for discount in range(len(EmpDisc['Date'][Bank])):
                                        if Date == pd.to_datetime(EmpDisc['Date'][Bank].iat[discount]):
                                            EmpDayTotal += EmpDisc['Item Subtotal'][Bank].iat[discount]
                                            cell_addition(text=EmpDayTotal,
                                                          working_cell=Trailing_Columns['Total Employee_1'] + Row,
                                                          font=Normal, number_format=Currency)
                                except TypeError:
                                    if Date.day == pd.to_datetime(EmpDisc['Date'][Bank]).day:
                                        cell_addition(text=EmpDisc['Item Subtotal'][Bank],
                                                      working_cell=Trailing_Columns['Total Employee_1'] + Row,
                                                      font=Normal, number_format=Currency)
                                    break
                            except KeyError:
                                pass
                    else:
                        # PAY PERIODS
                        elseSC = data_frame_try_catch(df=Tendered, group='SCTotal', location=Bank, place=data)
                        SC_COM = round(elseSC / (CurrentTax + 1), 2)
                        SCTax = round((elseSC - SC_COM), 2)
                        # Tax Calculations
                        elseSCTaxed = data_frame_try_catch(df=Tendered,
                                                           group='SCTotal Taxed', location=Bank, place=data)
                        if np.isnan(elseSCTaxed):
                            elseGCTaxed = data_frame_try_catch(df=Tendered,
                                                               group='GTotal Taxed', location=Bank, place=data)
                            if elseSC < 0:

                                cell_addition(text=(round(elseGCTaxed, 2)
                                                    - SCTax), working_cell=Totals_Columns['Tax Included'] + Row,
                                              font=Normal, number_format=Currency)
                            else:  # Pretty sure these do the same thing...
                                cell_addition(text=(round(elseGCTaxed, 2)
                                                    - SCTax), working_cell=Totals_Columns['Tax Included'] + Row,
                                              font=Normal, number_format=Currency)

                        else:
                            elseGTTaxed = data_frame_try_catch(df=Tendered, group='GTotal Taxed',
                                                               location=Bank, place=data)
                            if elseSC < 0:
                                cell_addition(text=(round(elseGTTaxed, 2)
                                                    - SCTax), working_cell=Totals_Columns['Tax Included'] + Row,
                                              font=Normal, number_format=Currency)
                            else:
                                # SCTax += abs(Tendered['SCTotal Taxed'][Bank].iat[data])
                                cell_addition(text=(round(elseGTTaxed, 2)
                                                    - SCTax), working_cell=Totals_Columns['Tax Included'] + Row,
                                              font=Normal, number_format=Currency)

                        if Date.day <= 15:
                            # PAY PERIODS
                            # This Try Block adds the Employee payments
                            try:
                                try:
                                    EmpDayTotal = 0
                                    for discount in range(len(EmpDisc['Date'][Bank])):
                                        fifEmpDate = data_frame_try_catch(df=EmpDisc, group='Date',
                                                                          location=Bank, place=discount)
                                        if Date.day == pd.to_datetime(fifEmpDate).day:
                                            EmpDayTotal += data_frame_try_catch(df=EmpDisc, group='Item Subtotal',
                                                                                location=Bank, place=discount)
                                            cell_addition(text=EmpDayTotal,
                                                          working_cell=Trailing_Columns['Total Employee_2'] + Row,
                                                          font=Normal, number_format=Currency)
                                except TypeError:
                                    if Date.day == pd.to_datetime(EmpDisc['Date'][Bank]).day:
                                        cell_addition(text=(EmpDisc['Item Subtotal'][Bank]),
                                                      working_cell=Trailing_Columns['Total Employee_2'] + Row,
                                                      font=Normal, number_format=Currency)
                                    break
                            except KeyError:
                                pass
                        else:
                            cell_addition(
                                text=('={1}{0}-{2}{0}-{3}{0}+{4}{0}'.format(Row, Totals_Columns['Grand Total'],
                                                                            Totals_Columns['Tax Included'],
                                                                            Trailing_Columns[
                                                                                'Total Employee_1'],
                                                                            Trailing_Columns['GC bought'])),
                                working_cell=Trailing_Columns['PPS2'] + Row,
                                font=Normal, number_format=Currency)
                            # This Try Block adds the Employee payments
                            try:
                                try:
                                    EmpDayTotal = 0
                                    for discount in range(len(EmpDisc['Date'][Bank])):
                                        if Date == pd.to_datetime(EmpDisc['Date'][Bank].iat[discount]):
                                            EmpDayTotal += EmpDisc['Item Subtotal'][Bank].iat[discount]
                                            cell_addition(text=EmpDayTotal,
                                                          working_cell=Trailing_Columns['Total Employee_1'] + Row,
                                                          font=Normal, number_format=Currency)
                                except TypeError:
                                    if Date.day == pd.to_datetime(EmpDisc['Date'][Bank]).day:
                                        cell_addition(text=EmpDisc['Item Subtotal'][Bank],
                                                      working_cell=Trailing_Columns['Total Employee_1'] + Row,
                                                      font=Normal, number_format=Currency)
                            except KeyError:
                                pass
    # Full month's total calculations
    CellValue += 1
    ColumnSummationRange = (int(Row) - FullMonth, CellValue - 1)
    cell_addition(text=(Locations_Key[str(Bank)].split(' ')[0] + ' Total'), working_cell=('A' + str(CellValue)),
                  font=Bold, number_format=Currency)  # Total

    cell_addition(text=('=SUM({0}{1}:{0}{2})'.format(Cash_Columns['Nadeau Reports'], ColumnSummationRange[0],
                                                     ColumnSummationRange[1])),
                  working_cell=('{0}{1}'.format(Cash_Columns['Nadeau Reports'], CellValue)), font=Normal,
                  number_format=Currency)  # Cash
    cell_addition(text=('=SUM({0}{1}:{0}{2})'.format(Cash_Columns['Per Statements'], ColumnSummationRange[0],
                                                     ColumnSummationRange[1])),
                  working_cell=('{0}{1}'.format(Cash_Columns['Per Statements'], CellValue)), font=Normal,
                  number_format=Currency)
    cell_addition(text=('={1}{0}-{2}{0}'.format(CellValue, Cash_Columns['Per Statements'],
                                                Cash_Columns['Nadeau Reports'])),
                  working_cell=('{0}{1}'.format(Cash_Columns['Difference'], CellValue)), font=Normal,
                  number_format=Currency)

    cell_addition(text=('=SUM({0}{1}:{0}{2})'.format(Check_Columns['Nadeau Reports'], ColumnSummationRange[0],
                                                     ColumnSummationRange[1])),
                  working_cell=('{0}{1}'.format(Check_Columns['Nadeau Reports'], CellValue)), font=Normal,
                  number_format=Currency)  # Check
    cell_addition(text=('=SUM({0}{1}:{0}{2})'.format(Check_Columns['Per Statements'], ColumnSummationRange[0],
                                                     ColumnSummationRange[1])),
                  working_cell=('{0}{1}'.format(Check_Columns['Per Statements'], CellValue)), font=Normal,
                  number_format=Currency)
    cell_addition(text=('={1}{0}-{2}{0}'.format(CellValue, Check_Columns['Per Statements'],
                                                Check_Columns['Nadeau Reports'])),
                  working_cell=('{0}{1}'.format(Check_Columns['Difference'], CellValue)), font=Normal,
                  number_format=Currency)

    cell_addition(text=('=SUM({0}{1}:{0}{2})'.format(VisaMCD_Columns['Nadeau Reports'], ColumnSummationRange[0],
                                                     ColumnSummationRange[1])),
                  working_cell=('{0}{1}'.format(VisaMCD_Columns['Nadeau Reports'], CellValue)), font=Normal,
                  number_format=Currency)  # Visa
    cell_addition(text=('=SUM({0}{1}:{0}{2})'.format(VisaMCD_Columns['Per Statements'], ColumnSummationRange[0],
                                                     ColumnSummationRange[1])),
                  working_cell=('{0}{1}'.format(VisaMCD_Columns['Per Statements'], CellValue)), font=Normal,
                  number_format=Currency)
    cell_addition(text=('={1}{0}-{2}{0}'.format(CellValue, VisaMCD_Columns['Per Statements'],
                                                VisaMCD_Columns['Nadeau Reports'])),
                  working_cell=('{0}{1}'.format(VisaMCD_Columns['Difference'], CellValue)), font=Normal,
                  number_format=Currency)

    cell_addition(text=('=SUM({0}{1}:{0}{2})'.format(Amex_Columns['Nadeau Reports'], ColumnSummationRange[0],
                                                     ColumnSummationRange[1])),
                  working_cell=('{0}{1}'.format(Amex_Columns['Nadeau Reports'], CellValue)), font=Normal,
                  number_format=Currency)  # AMEX
    cell_addition(text=('=SUM({0}{1}:{0}{2})'.format(Amex_Columns['Per Statements'], ColumnSummationRange[0],
                                                     ColumnSummationRange[1])),
                  working_cell=('{0}{1}'.format(Amex_Columns['Per Statements'], CellValue)), font=Normal,
                  number_format=Currency)
    cell_addition(text=('={1}{0}-{2}{0}'.format(CellValue, Amex_Columns['Per Statements'],
                                                Amex_Columns['Nadeau Reports'])),
                  working_cell=('{0}{1}'.format(Amex_Columns['Difference'], CellValue)), font=Normal,
                  number_format=Currency)

    cell_addition(text=('=SUM({0}{1}:{0}{2})'.format(Square_Columns['Nadeau Reports'], ColumnSummationRange[0],
                                                     ColumnSummationRange[1])),
                  working_cell=('{0}{1}'.format(Square_Columns['Nadeau Reports'], CellValue)), font=Normal,
                  number_format=Currency)
    cell_addition(text=('=SUM({0}{1}:{0}{2})'.format(Square_Columns['Per Statements'], ColumnSummationRange[0],
                                                     ColumnSummationRange[1])),
                  working_cell=('{0}{1}'.format(Square_Columns['Per Statements'], CellValue)), font=Normal,
                  number_format=Currency)
    cell_addition(text=('={1}{0}-{2}{0}'.format(CellValue, Square_Columns['Per Statements'],
                                                Square_Columns['Nadeau Reports'])),
                  working_cell=('{0}{1}'.format(Square_Columns['Difference'], CellValue)), font=Normal,
                  number_format=Currency)

    cell_addition(text=('=SUM({0}{1}:{0}{2})'.format(Totals_Columns['Credit Card'], ColumnSummationRange[0],
                                                     ColumnSummationRange[1])),
                  working_cell=('{0}{1}'.format(Totals_Columns['Credit Card'], CellValue)), font=Normal,
                  number_format=Currency)
    cell_addition(text=('=SUM({0}{1}:{0}{2})'.format(Totals_Columns['Grand Total'], ColumnSummationRange[0],
                                                     ColumnSummationRange[1])),
                  working_cell=('{0}{1}'.format(Totals_Columns['Grand Total'], CellValue)), font=Normal,
                  number_format=Currency)
    cell_addition(text=('=SUM({0}{1}:{0}{2})'.format(Totals_Columns['GC Used'], ColumnSummationRange[0],
                                                     ColumnSummationRange[1])),
                  working_cell=('{0}{1}'.format(Totals_Columns['GC Used'], CellValue)), font=Normal,
                  number_format=Currency)
    cell_addition(text=('=SUM({0}{1}:{0}{2})'.format(Totals_Columns['Tax Included'], ColumnSummationRange[0],
                                                     ColumnSummationRange[1])),
                  working_cell=('{0}{1}'.format(Totals_Columns['Tax Included'], CellValue)), font=Normal,
                  number_format=Currency)

    cell_addition(text=('=SUM({0}{1}:{0}{2})'.format(Trailing_Columns['GC bought'], ColumnSummationRange[0],
                                                     ColumnSummationRange[1])),
                  working_cell=('{0}{1}'.format(Trailing_Columns['GC bought'], CellValue)), font=Normal,
                  number_format=Currency)
    cell_addition(text=('=SUM({0}{1}:{0}{2})'.format(Trailing_Columns['PPS2'], ColumnSummationRange[0],
                                                     ColumnSummationRange[1])),
                  working_cell=('{0}{1}'.format(Trailing_Columns['PPS2'], CellValue)), font=Normal,
                  number_format=Currency)
    cell_addition(text=('=SUM({0}{1}:{0}{2})'.format(Trailing_Columns['Total Employee_1'], ColumnSummationRange[0],
                                                     ColumnSummationRange[1])),
                  working_cell=('{0}{1}'.format(Trailing_Columns['Total Employee_1'], CellValue)),
                  font=Normal, number_format=Currency)

    cell_addition(text=('=SUM({0}{1}:{0}{2})'.format(Trailing_Columns['PP1'], ColumnSummationRange[0],
                                                     ColumnSummationRange[1])),
                  working_cell=('{0}{1}'.format(Trailing_Columns['PP1'], CellValue)),
                  font=Normal, number_format=Currency)
    cell_addition(text=('=SUM({0}{1}:{0}{2})'.format(Trailing_Columns['Total Employee_2'], ColumnSummationRange[0],
                                                     ColumnSummationRange[1])),
                  working_cell=('{0}{1}'.format(Trailing_Columns['Total Employee_2'], CellValue)),
                  font=Normal, number_format=Currency)
    cell_addition(text=('=SUM({0}{1}:{0}{2})'.format(Trailing_Columns['Difference_2'], ColumnSummationRange[0],
                                                     ColumnSummationRange[1])),
                  working_cell=('{0}{1}'.format(Trailing_Columns['Difference_2'], CellValue)),
                  font=Normal, number_format=Currency)

    RED = 'AA0000'
    babyBLUE = '00ABFF'
    YELLOW = 'FFFF00'
    # CurrentTax = globals().get('CurrentTax')
    if CurrentTax > 0:
        rule = ColorScaleRule(start_type='num', start_value=CurrentTax - 0.01, start_color=RED,
                              mid_type='num', mid_value=CurrentTax, mid_color=babyBLUE,
                              end_type='num', end_value=CurrentTax + 0.01, end_color=YELLOW)
    else:
        rule = ColorScaleRule(start_type='num', start_value=0.0000, start_color=babyBLUE,
                              mid_type='num', mid_value=CurrentTax, mid_color=babyBLUE,
                              end_type='num', end_value=CurrentTax + 0.01, end_color=YELLOW)

    # dxf = DifferentialStyle(fill=RED)

    work_sheet.conditional_formatting.add(('{0}{1}:{0}{2}'.format(Trailing_Columns['Verification'],
                                                                  int(Row) - FullMonth, CellValue - 1)), rule)

    for cell in work_sheet[CellValue:CellValue]:
        cell.border = RowBorderSeparator

    CellValue += 1

if __name__ == '__main__':
    worksheet.dimensions.ColumnDimension(work_sheet, bestFit=True)
    wb.save(filename=filename + '.xlsx')
